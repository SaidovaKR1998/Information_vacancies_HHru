import csv
import json
from abc import ABC, abstractmethod
from typing import List, Dict, Any
import openpyxl  # pip install openpyxl


class CSVSaver:
    """Класс для сохранения вакансий в CSV"""

    def __init__(self, filename: str = 'data/vacancies.csv'):
        self.__filename = filename

    def add_vacancy(self, vacancy) -> None:
        vacancies = self._load_vacancies()

        if not any(v['url'] == vacancy.url for v in vacancies):
            vacancies.append(vacancy.__dict__)
            self._save_vacancies(vacancies)

    def _load_vacancies(self) -> List[Dict[str, Any]]:
        try:
            with open(self.__filename, mode='r', encoding='utf-8') as file:
                return list(csv.DictReader(file))
        except FileNotFoundError:
            return []

    def _save_vacancies(self, vacancies: List[Dict[str, Any]]) -> None:
        with open(self.__filename, mode='w', encoding='utf-8', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=vacancies[0].keys())
            writer.writeheader()
            writer.writerows(vacancies)


class ExcelSaver:
    """Класс для сохранения вакансий в Excel"""

    def __init__(self, filename: str = 'data/vacancies.xlsx'):
        self.__filename = filename

    def add_vacancy(self, vacancy) -> None:
        vacancies = self._load_vacancies()

        if not any(v['url'] == vacancy.url for v in vacancies):
            vacancies.append(vacancy.__dict__)
            self._save_vacancies(vacancies)

    def _load_vacancies(self) -> List[Dict[str, Any]]:
        try:
            workbook = openpyxl.load_workbook(self.__filename)
            sheet = workbook.active
            return [
                dict(zip([cell.value for cell in sheet[1]], row))
                for row in sheet.iter_rows(min_row=2, values_only=True)
            ]
        except FileNotFoundError:
            return []

    def _save_vacancies(self, vacancies: List[Dict[str, Any]]) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        if vacancies:
            sheet.append(list(vacancies[0].keys()))
            for vacancy in vacancies:
                sheet.append(list(vacancy.values()))

        workbook.save(self.__filename)
